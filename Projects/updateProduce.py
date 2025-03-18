#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
from pathlib import Path

# Get the current  path
currentPath = Path(__file__).parent.parent

wb = openpyxl.load_workbook(currentPath / 'ExcelFiles' / 'produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

print('Updating Prices...')
# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save(currentPath / 'ExcelFiles' / 'updatedProduceSales.xlsx')

print('Done.')
